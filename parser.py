import openpyxl

def handle_uploaded_file():
    # wb = openpyxl.load_workbook(f)
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['Sheet1']

    part_list = []
    for row in range(2, ws.max_row + 1):
        x = {}
        if ws['A' + str(row)].font.color == "#110000" or ws['A' + str(row)].font.color == "#100":
            x['status'] = 3
        elif ws['A' + str(row)].font.color == "#111100" or ws['A' + str(row)].font.color == "#110":
            x['status'] = 2
        else:
            x['status'] = 1
        
        x['reported_on'] = ws['B' + str(row)].value
        x['short_on'] = ws['C' + str(row)].value
        x['shop'] = ws['D' + str(row)].value
        x['variants'] = ws['E' + str(row)].value
        # x['count'] = ws['' + str(row)].value
        x['part_number'] = ws['G' + str(row)].value
        x['description'] = ws['H' + str(row)].value
        x['supplier_name'] = ws['I' + str(row)].value
        x['pmc'] = ws['J' + str(row)].value
        x['team'] = ws['K' + str(row)].value
        x['backlog'] = ws['N' + str(row)].value
        x['region'] = ws['P' + str(row)].value
        x['unloading_point'] = ws['Q' + str(row)].value
        x['p_q'] = ws['R' + str(row)].value
        x['quantity'] = ws['S' + str(row)].value
        x['quantity_expected'] = ws['T' + str(row)].value
        x['planned_vehicle_qty'] = ws['U' + str(row)].value
        x['eta_dicv'] = ws['V' + str(row)].value
        x['truck_details'] = ws['W' + str(row)].value
        x['shortage_reason'] = ws['X' + str(row)].value
                
        part_list.append(x)

    # convertToDB(part_list)
    return part_list


print handle_uploaded_file()