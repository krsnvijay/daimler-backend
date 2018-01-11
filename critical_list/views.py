import datetime
from threading import Thread

import openpyxl
from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse, Http404
from django.shortcuts import render
# Create your views here.
from django_filters import rest_framework as filters
from django_filters.rest_framework import DjangoFilterBackend
from oauth2_provider.contrib.rest_framework import OAuth2Authentication, IsAuthenticatedOrTokenHasScope
from pywebpush import webpush, WebPushException
from rest_framework import viewsets, status
from rest_framework.authentication import TokenAuthentication, SessionAuthentication
from rest_framework.filters import OrderingFilter
from rest_framework.permissions import DjangoModelPermissions
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.models import Subscription
from critical_list.forms import UploadFileForm
from critical_list.models import Part
from critical_list.permissions import IsManagerOrReadOnly
from critical_list.serailizers import PartSerializer
from sos.serializers import PartNotificationSerializer


def get_starred_parts(request):
    user = request.user
    return user.starred_parts.all()


class PartFilter(filters.FilterSet):
    pmcvalues = (
        ('Arulselvan', 'HDT ENGINE'),
        ('Balaji', 'TRANSMISSION'),
        ('Joshna', 'AXLE'),
        ('Giftson', 'MDT ENGINE'),
        ('Premkumar', 'CASTING AND FORGING'),
    )
    daterange = filters.DateFromToRangeFilter(name="short_on")
    pmc = filters.ChoiceFilter(choices=pmcvalues)

    class Meta:
        model = Part
        fields = ['shop', 'status', 'daterange', 'short_on', 'pmc']


class PartViewSet(viewsets.ModelViewSet):
    authentication_classes = (TokenAuthentication, OAuth2Authentication, SessionAuthentication)
    permission_classes = [IsAuthenticatedOrTokenHasScope, DjangoModelPermissions]
    queryset = Part.objects.all()
    serializer_class = PartSerializer
    filter_backends = (DjangoFilterBackend, OrderingFilter)
    search_fields = ('part_number',)
    ordering_fields = ('shop', 'short_on', 'status', 'pmc')
    filter_class = PartFilter



def convertToDB(records):
    for record in records:
        entry = Part()
        entry.part_number = record['part_number']
        entry.description = record['description']
        entry.supplier_name = record['supplier_name']
        entry.variants = record['variants']
        # entry.count = record['count']
        entry.reported_on = record['reported_on']
        entry.short_on = record['short_on']
        entry.shop = record['shop']
        entry.pmc = record['pmc']
        entry.team = record['team']
        entry.backlog = record['backlog']
        entry.region = record['region']
        entry.unloading_point = record['unloading_point']
        entry.p_q = record['p_q']
        entry.quantity = record['quantity']
        entry.quantity_expected = record['quantity_expected']
        entry.planned_vehicle_qty = record['planned_vehicle_qty']
        entry.eta_dicv = record['eta_dicv']
        entry.truck_details = record['truck_details']
        entry.shortage_reason = record['shortage_reason']
        entry.status = record['status']
        entry.save()


def handle_uploaded_file(url):
    # print (f.temporary_file_path())
    wb = openpyxl.load_workbook(open("." + url, "rb"))
    ws = wb['Dataset']

    part_list = []
    for row in range(3, ws.max_row + 1):
        x = {}
        if ws['A' + str(row)].font.color.rgb == "FFFF0000":
            x['status'] = 3
        # elif ws['A' + str(row)].font.color == "#111100" or ws['A' + str(row)].font.color == "#110":
        #     x['status'] = 2
        else:
            x['status'] = 1

        x['reported_on'] = ws['B' + str(row)].value
        x['short_on'] = ws['C' + str(row)].value
        x['shop'] = ws['D' + str(row)].value
        x['variants'] = ws['E' + str(row)].value
        # x['count'] = ws['' + str(row)].value
        x['part_number'] = ws['G' + str(row)].value
        x['description'] = ws['H' + str(row)].value
        x['supplier_name'] = ws['I' + str(row)].value
        x['pmc'] = ws['J' + str(row)].value
        x['team'] = ws['K' + str(row)].value
        x['backlog'] = ws['N' + str(row)].value
        x['region'] = ws['P' + str(row)].value
        x['unloading_point'] = ws['Q' + str(row)].value
        x['p_q'] = ws['R' + str(row)].value
        x['quantity'] = ws['S' + str(row)].value
        x['quantity_expected'] = ws['T' + str(row)].value
        x['planned_vehicle_qty'] = ws['U' + str(row)].value
        x['eta_dicv'] = ws['V' + str(row)].value
        x['truck_details'] = ws['W' + str(row)].value
        x['shortage_reason'] = ws['X' + str(row)].value

        part_list.append(x)

    # convertToDB(part_list)
    return part_list


def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            fs = FileSystemStorage()
            filename = fs.save('uploaded_sheets/' + request.FILES['file'].name, request.FILES['file'])
            uploaded_file_url = fs.url(filename)
            thread = Thread(target=handle_uploaded_file, args=[uploaded_file_url])
            thread.start()
            return JsonResponse({
                "Status": 200,
                "message": "Successfully uploaded the file",
                "file_url": uploaded_file_url
            }, safe=False)
    else:
        form = UploadFileForm()
    return render(request, 'upload.html', {'form': form})


class PartStatusChangeViewSet(APIView):
    authentication_classes = (TokenAuthentication, OAuth2Authentication, SessionAuthentication)
    permission_classes = [IsAuthenticatedOrTokenHasScope, IsManagerOrReadOnly]

    def patch(self, request, pk, format=None):
        try:
            obj = Part.objects.get(pk=pk)
            serializer = PartSerializer(obj, data=request.data, context={'request': request}, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            else:
                return Response(serializer.errors)
        except Part.DoesNotExist:
            raise Http404


class CriticalListViewSet(APIView):
    """
    get: Get Shoptypes and their part info
    """
    authentication_classes = (TokenAuthentication, OAuth2Authentication, SessionAuthentication)
    permission_classes = [IsAuthenticatedOrTokenHasScope]

    def get(self, request, format=None):
        values = {
            'Arulselvan': 'HDT ENGINE',
            'Balaji': 'TRANSMISSION',
            'Joshna': 'AXLE',
            'Giftson': 'MDT ENGINE',
            'Premkumar': 'CASTING AND FORGING',
        }
        q = Part.objects.all()
        try:
            date = self.request.query_params.get('short_on', datetime.datetime.today())
            x = {}
            x['total'] = q.filter(short_on=date).count()
            x['critical'] = q.filter(short_on=date, status=3).count()
            x['warning'] = q.filter(short_on=date, status=2).count()
            x['normal'] = q.filter(short_on=date, status=1).count()
            for shop in values.keys():
                pmc = values[shop]
                x[pmc] = {}
                o = {}
                o['pmc'] = shop
                o['total'] = q.filter(short_on=date, pmc=shop).count()
                o['critical'] = q.filter(short_on=date, pmc=shop, status=3).count()
                o['warning'] = q.filter(short_on=date, pmc=shop, status=2).count()
                o['normal'] = q.filter(short_on=date, pmc=shop, status=1).count()
                x[pmc] = o
            return Response(x)
        except ValueError:
            return Response({'detail': 'Invalid date format should be of form YYYY-MM-DD'},
                            status=status.HTTP_400_BAD_REQUEST)


class CriticalPartsViewSet(APIView):
    """
    get: Get Sorted Critical list starred>status
    """
    authentication_classes = (TokenAuthentication, OAuth2Authentication, SessionAuthentication)
    permission_classes = [IsAuthenticatedOrTokenHasScope]

    def get(self, request, format=None):
        q = Part.objects.all()
        try:
            date = self.request.query_params.get('short_on', datetime.datetime.today())
            parts = set(
                list(request.user.starred_parts.filter(short_on=date)) + list(q.filter(short_on=date, status=3)))
            serializer = PartSerializer(parts, many=True, context={'request': request})
            return Response(sorted(serializer.data, key=lambda item: item['starred'], reverse=True))
        except ValueError:
            return Response({'detail': 'Invalid date format should be of form YYYY-MM-DD'},
                            status=status.HTTP_400_BAD_REQUEST)


class PartNotificationViewSet(APIView):
    """
    post: Send Notification and star a part to a particular user(userid,partid,content required)
    """
    authentication_classes = (TokenAuthentication, OAuth2Authentication, SessionAuthentication)
    permission_classes = [IsAuthenticatedOrTokenHasScope]

    def post(self, request, format=None):
        serializer = PartNotificationSerializer(data=request.data, context={'request': request})
        if (serializer.is_valid()):
            comment = serializer.save()
            user = comment.userid
            print(comment.partid)
            user.starred_parts.add(comment.partid)
            user.save()
            # TODO implement webpush
            try:
                subscription = Subscription.objects.get(user_id=user)
                webpush(subscription_info=subscription.subscription_id,
                        data="You've been notified of part {part}-{shop}-{status} by {poster}".format(
                            part=comment.partid.part_number, poster=comment.posted_by.username,
                            shop=comment.partid.shop, status=Part.status_values[comment.partid.status]),
                        vapid_private_key="5tOi9dKR77pqY0uQ5H2PqQbR6YMG1c75A2XgR7izOcA",
                        vapid_claims={
                            "sub": "mailto:YourNameHere@example.org",
                        })
            except WebPushException as ex:
                print("Push Notification not sent {}", repr(ex))

            return Response({'Notification Sent to ' + user.first_name + ' about part ' + comment.partid.part_number},
                            status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
